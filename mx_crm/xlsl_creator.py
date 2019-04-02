from openpyxl import Workbook

from mx_crm import settings
from mx_crm.calculation.location import LocationEvaluationLevel

# book = Workbook()
# book.create_sheet('report')

from mx_crm.settings.base import RATING_HEADERS

companies = [u'Fiducia und GAD IT AG, Karlsruhe', u'CARL ZEISS AG', u'rhoen-klinikum-ag', u'mvv ag', u'testo ag', u'Sanitaetshaus Mueller Betten', u'tkk', u'carl beutlhauser baumaschinen gmbh', u'Essener Verkehrs-AG', u'badenit gmbh, managed services rz network', u'fls gmbh', u'envia Mitteldeutsche energie AG', u'ncc-steyr', u'Stadtwerke Hamm', u'ITELLIGENCE AG', u'htp-maxschoen', u'Post CH AG']


# wb_headers = RATING_HEADERS
# ws = book.create_sheet('report')

loc_l = LocationEvaluationLevel().calc(companies=companies)

print loc_l



