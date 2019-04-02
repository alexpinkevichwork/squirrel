# -*- coding: utf-8 -*-
from __future__ import print_function, absolute_import

import os
import re
import logging
import datetime
from pprint import pprint

from openpyxl import Workbook

from mx_crm import settings
from mx_crm.calculation.branch import BranchEvaluationLevel
from mx_crm.calculation.google import GoogleEvaluationLevel
from mx_crm.calculation.location import LocationEvaluationLevel
from mx_crm.calculation.revenue_size import RevenueSizeEvaluationLevel
from mx_crm.calculation.squirrel_rating import SquirrelRating
from mx_crm.queries import get_google_analytics_sessions
from mx_crm.utils import prepare_company_name_for_match, convert_to_int, convert_to_float, \
    add_google_analytics_accounts_to_report_file
from mx_crm.models import session as alchemy_session, WikipediaDb, XingCompanyDb, Company, session, MxCrmAccessHistory

logger = logging.getLogger(__name__)


def get_companies_info(companies):
    companies_entries = alchemy_session.query(Company). \
        filter(Company.name.in_(companies) & (Company.website != None) & (Company.website != 'NA'))
    return {c.name.lower(): c for c in companies_entries}


def get_companies_info_websites(companies):
    companies_entries = alchemy_session.query(Company). \
        filter(Company.name.in_(companies) & (Company.website != None) & (Company.website != 'NA'))
    return {c.website.lower(): c for c in companies_entries}


def get_wiki_info(companies):
    wiki_entries = alchemy_session.query(WikipediaDb).filter(WikipediaDb.company_name_w.in_(companies))
    return {c.company_name_w.lower(): c for c in wiki_entries}


def get_company_table_info(companies):
    company_entries = alchemy_session.query(Company).filter(Company.name.in_(companies))
    return {c.name.lower(): c for c in company_entries}


def get_xing_info(companies):
    xing_entries = alchemy_session.query(XingCompanyDb).join(Company, XingCompanyDb.xc_id == Company.id). \
        filter(Company.name.in_(companies), Company.xing_page is not None, Company.xing_page != 'NA')
    return {c.company_name_x.lower(): c for c in xing_entries}


def get_manual_website(name):
    query = session.query(Company.name, Company.manual_account_id).filter(Company.name == name)
    try:
        return {name: query[0][1]}
    except IndexError:
        return {name: u'No'}


def get_manual_account(company_names):
    returned_obj = {}
    for name in company_names:
        obj = {}
        query = session.query(Company.name, Company.manual_account_id).filter(Company.name == name)
        try:
            status = u'{}'.format(query[0][1])
            obj[name] = status
            returned_obj.update(obj)
        except IndexError:
            status = u'Not in db'
            obj[name] = status
            returned_obj.update(obj)
    return returned_obj


def _get_and_prepare_company_names_from_objects(self, companie_entries):
    return set(
        company.name for company in companie_entries
    )


def create_report(companies, account_data=[], account_headers=[], total_fields=[], data_links={}, google_analytics_companies={},
                  dates={}):
    """
    Creates and saves locally report.
    :param companies: List of companies that made requests during specified range
    """
    logger.debug(companies)

    file_name = settings.REPORTS_FILE.format(now=datetime.datetime.now().strftime("%Y_%m_%d-%H_%M_%S"))
    path_to_xl = settings.rel('mx_crm', settings.REPORTS_FOLDER, file_name)
    logger.debug('Export excel file: {}'.format(path_to_xl))

    wb = Workbook()
    ws = wb.create_sheet('Report')

    logger.info('Saving report to the local excel file')
    wb_headers = settings.NEW_WORKBOOK_HEADERS
    # wb_headers = settings.WORKBOOK_HEADERS
    if account_headers:
        wb_headers += account_headers
    if total_fields:
        wb_headers += settings.TOTAL_HEADERS
    wb_headers += settings.RATING_HEADERS
    ws.append(wb_headers)

    companies_info = get_companies_info(companies)
    logger.info('companies_info')
    logger.info(companies_info)
    companies_info_manual_id = get_company_table_info(companies)
    logger.info('companies_info_manual_id')
    logger.info(companies_info_manual_id)
    # manual
    companies_info_websites = get_companies_info_websites(companies)
    logger.debug('Companies: {}'.format(len(companies_info)))

    companies_wiki_info = get_wiki_info(companies)
    logger.debug('Wiki companies: {}'.format(len(companies_wiki_info)))

    companies_xing_info = get_xing_info(companies)
    logger.debug('Xing companies: {}'.format(len(companies_xing_info)))

    companies_names = set()
    websites_for_rating = set()
    for c in companies_info.values():
        if c.website:
            websites_for_rating.add(c.website)
        if c.name:
            companies_names.add(c.name)
    rating_data = SquirrelRating().calc(companies=companies_names, websites=websites_for_rating)
    company_manual_account = get_manual_account(companies_names)
    variables_data = SquirrelRating().get_rating_variables(companies, websites_for_rating)
    #logger.info("rating data {}".format(rating_data))
    #logger.info("rating data {}".format(type(rating_data)))

    try:
        counter = 0
        for company_name, company in sorted(companies.items(), key=lambda x: x[1].session_length, reverse=True):
            ws.row_dimensions[counter].collapsed = True
            address = company.full_address
            country = company.country
            # rating = rating_data.get(company.company_name).get('score')
            wiki_info = companies_wiki_info.get(company_name)
            xing_info = companies_xing_info.get(company_name)
            company_info = companies_info.get(company_name)
            company_table_manual_id = companies_info_manual_id.get(company_name)
            website = company_info.website if company_info else ''
            full_website = re.sub('www\d?\.', '', website).rstrip('/').lower()

            prepared_company_name = company_name
            xing_page = company_info.xing_page if company_info else None
            session_length = company.session_length

            for session in company.sessions:
                for request in session.requests:
                    #master_company = alchemy_session.query(Company.name).filter(Company.name == company.company_name)
                    access_history = MxCrmAccessHistory(
                        company_name=company.company_name,
                        a_h_sid=counter,
                        mx_crm_visited_page=request.title,
                        mx_crm_referrer=request.url[:255],
                        mx_crm_session_date=datetime.datetime.fromtimestamp(int(request.timestamp)).strftime(
                            '%Y-%m-%d'),
                        mx_crm_session_time=datetime.datetime.fromtimestamp(int(request.timestamp)).strftime(
                            '%H:%M:%S'),
                        mx_crm_ip_vlan=request.hostname
                    )
                    alchemy_session.add(access_history)
                    alchemy_session.commit()
                    sheet_counter = 2
                    company_table_info = get_manual_website(company.company_name)
                    access_dt = datetime.datetime.fromtimestamp(request.timestamp).strftime('%Y-%m-%d %H:%M:%S')
                    rcd_name_rating = companies_info.get(company_name)
                    if rcd_name_rating and rcd_name_rating.name:
                        rating = rating_data.get(rcd_name_rating.name, 'N/C') if rating_data.get(
                            rcd_name_rating.name) is not None else 'N/C'

                    if company_name in total_fields:
                        obj = total_fields.get(company_name, {})
                        total_session_lenght = datetime.timedelta(seconds=obj.get('time') or 0)
                    # row = [company.company_name]
                    sheet_number = 'A{}'.format(sheet_counter)
                    # ws[sheet_number].hyperlink = "http://google.com"
                    # ws[sheet_number].value = company.company_name
                    # ws.cell(row=1, column=sheet_counter).value = '=HYPERLINK("{}", "{}")'.format('google.com', company.company_name)
                    link = ''
                    # pprint(company.company_name)
                    link = data_links.get(company.company_name.lower())
                    c_id = alchemy_session.query(Company.id).filter(Company.name == company.company_name)

                    try:
                        company_id = c_id[0][0]
                        webinterface_link = "http://192.168.0.141:8000/squirrel/accounts/{}/".format(company_id)
                    except IndexError:
                        company_id = ''
                    webinterface_link = "http://192.168.0.141:8000/squirrel/accounts/search/{}/".format(company.company_name)
                    # pprint(link)

                    query_link = alchemy_session.query(Company).filter(Company.name == company.company_name)
                    query_link.update({Company.d_crm_link: link}, synchronize_session="fetch")
                    alchemy_session.commit()
                    row = ['=HYPERLINK("{}", "{}")'.format(webinterface_link, company.company_name),
                           company_table_info.get(company.company_name), website, session_length, total_session_lenght,
                           rating_data.get(company.company_name), address, request.title,
                           request.url, access_dt, country]
                    sheet_counter += 1
                    # pprint(type(row))
                    if wiki_info:
                        row.extend([
                            wiki_info.manual_entry,
                            wiki_info.wiki_url_w,
                            convert_to_float(wiki_info.revenue_wikipedia_w),
                            wiki_info.revenue_currency_wiki_w,
                            convert_to_int(wiki_info.employees_wikipedia_w),
                            wiki_info.categories_wikipedia_w,
                            wiki_info.branch_wikipedia_w,
                            wiki_info.summary_wikipedia_w,
                        ])
                    else:
                        row.extend([''] * 8)

                    if xing_info:
                        if company_table_manual_id.manual_account_id:
                            c_t_manual_id = company_table_manual_id.manual_account_id
                        elif company_table_manual_id.manual_account_id == u'':
                            c_t_manual_id = u'NONE'
                        elif company_table_manual_id.manual_account_id == '':
                            c_t_manual_id = u'NONE'
                        else:
                            c_t_manual_id = u'NONE'
                        row.extend([
                            xing_info.manual_entry,
                            xing_page,
                            xing_info.country_xing,
                            xing_info.employees_group_xing_x,
                            xing_info.employees_size_xing,
                            xing_info.description_xing,
                            xing_info.industry_xing,
                            c_t_manual_id
                            # company_manual_account.get(company_name)
                        ])
                    else:
                        row.extend([''] * 8)

                    if full_website in account_data or prepared_company_name in account_data:
                        data_to_extend = []
                        for key in account_headers:
                            if full_website in account_data:
                                value = account_data[full_website].get(key, '')
                            else:
                                value = account_data[prepared_company_name].get(key, '')
                            data_to_extend.append(value)
                        row.extend(data_to_extend)
                    elif account_headers:
                        row.extend([''] * len(account_headers))

                    if company_name in total_fields:
                        obj = total_fields.get(company_name, {})
                        row.extend([
                            datetime.timedelta(seconds=obj.get('time') or 0),
                            convert_to_int(obj.get('visited')),
                            obj.get('last_visited'),
                        ])
                    else:
                        row.extend([''] * len(settings.TOTAL_HEADERS))

                    rcd_name = companies_info.get(company_name)
                    if rcd_name and rcd_name.name:
                        if wiki_info:
                            row.extend([
                                wiki_info.manual_entry
                            ])
                        else:
                            row.extend([""])

                        if xing_info:
                            row.extend([
                                xing_info.manual_entry
                            ])
                        else:
                            row.extend([""])
                        query = alchemy_session.query(Company).filter(Company.name == rcd_name.name)
                        dict_for_save = dict(mx_crm_location_level=variables_data.get(rcd_name.name).get('location'),
                                             mx_crm_branch_level=variables_data.get(rcd_name.name).get('branch'),
                                             mx_crm_google_evaluation=variables_data.get(rcd_name.name).get(
                                                 'google_ev'),
                                             mx_crm_wiki_rating_points=variables_data.get(rcd_name.name).get(
                                                 'wiki_size'),
                                             mx_crm_xing_rating_points=variables_data.get(rcd_name.name).get(
                                                 'xing_size'),
                                             mx_crm_revenue_level=variables_data.get(rcd_name.name).get(
                                                 'revenue_point'))
                        rating_update_info = dict(
                            mx_crm_location_level=variables_data.get(rcd_name.name).get('location'),
                            mx_crm_branch_level=variables_data.get(rcd_name.name).get('branch'),
                            mx_crm_google_evaluation=float(variables_data.get(rcd_name.name).get('google_ev')),
                            mx_crm_wiki_rating_points=variables_data.get(rcd_name.name).get('wiki_size'),
                            mx_crm_xing_rating_points=variables_data.get(rcd_name.name).get('xing_size'),
                            mx_crm_revenue_level=variables_data.get(rcd_name.name).get('revenue_point'))
                        query.update(rating_update_info, synchronize_session=False)
                        relation_ship_type = row[36]
                        account_name = row[27]
                        account_owner = row[28]
                        abc_rating = row[38]

                        closed_activity_type = row[31]
                        if row[32] != '':
                            closed_date = row[32]
                        else:
                            closed_date = None
                        # closed_date = datetime.datetime.strptime(str(row[32]), '%m/%d/%Y %H:%M:%S')
                        open_activity_type = row[33]
                        if row[34] != '':
                            schedule_date = row[34]
                        else:
                            schedule_date = None
                        # schedule_date = datetime.datetime.strptime(str(row[34]), '%m/%d/%Y %H:%M:%S')
                        total_session_length = row[39]
                        total_visited_page = row[40]
                        last_visit_time = row[41]

                        alchemy_session.commit()
                        dynamics_crm_info = dict(d_crm_relationship_type=relation_ship_type,
                                                 d_crm_account_name=account_name,
                                                 d_crm_account_owner=account_owner,
                                                 d_crm_abc_rating=abc_rating,
                                                 d_crm_closed_activity_type=closed_activity_type,
                                                 d_crm_open_activity_type=open_activity_type,
                                                 d_crm_closed_date=closed_date,
                                                 d_crm_schedule_date=schedule_date,
                                                 mx_crm_total_session_length=total_session_length,
                                                 mx_crm_total_visited_pages=total_visited_page,
                                                 mx_crm_last_visit=last_visit_time,
                                                 squirrel_rating=rating_data.get(rcd_name.name))
                                                 #webinterface_link=webinterface_link) # also in this query save webinterface link

                        query_dynamics_crm = alchemy_session.query(Company).filter(Company.name == rcd_name.name)
                        query_dynamics_crm.update(dynamics_crm_info, synchronize_session=False)
                        alchemy_session.commit()
                        row.extend([
                            rating_data.get(rcd_name.name, 'N/C') if rating_data.get(
                                rcd_name.name) is not None else 'N/C',
                        ])
                        row.extend([
                            variables_data.get(rcd_name.name).get('location')
                        ])
                        row.extend([
                            variables_data.get(rcd_name.name).get('branch')
                        ])
                        row.extend([
                            variables_data.get(rcd_name.name).get('google_ev')
                        ])
                        row.extend([
                            variables_data.get(rcd_name.name).get('wiki_size')
                        ])
                        row.extend([
                            variables_data.get(rcd_name.name).get('xing_size')
                        ])
                        row.extend([
                            variables_data.get(rcd_name.name).get('revenue_point')
                        ])
                    else:
                        row.extend(['N/C'] * len(settings.RATING_HEADERS))

                    try:
                        ws.append(row)
                    except ValueError as e:
                        logger.info(e)
                    counter += 1
                    if not ws.row_dimensions[counter - 1].collapsed:
                        ws.row_dimensions[counter].hidden = True
                        ws.row_dimensions[counter].outlineLevel = 1

        wb.save(path_to_xl)
        d_start = dates.get('start_date')
        e_date = dates.get('end_date')
        start_date = datetime.datetime(d_start.year, d_start.month, d_start.day)
        end_date = datetime.datetime(e_date.year, e_date.month, e_date.day)
        # g_a_c = get_google_analytics_sessions(start_date, end_date, True)
        # logger.info(g_a_c)
        # logger.info(google_analytics_companies)
        # result = add_google_analytics_accounts_to_report_file(path_to_xl, start_date, end_date)
        # os.chdir("C:/Users/admin/PycharmProjects/SquirrelRunnerNew/mx_crm")
        # cd = os.system('python add_companies.py --days_start={0} --year_start={1} --month_start={2} --days_end={3} --year_end={4} --month_end={5}'.format(
        #     d_start.day, d_start.year, d_start.month, e_date.day, e_date.year, e_date.month
        # ))
        # logger.info(cd)

    except KeyError as e:
        logger.error(e)
    logger.info('Local file has been updated')
