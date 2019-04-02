import re
import regex

from collections import OrderedDict

from openpyxl import load_workbook

from mx_crm import settings
from mx_crm.settings import RESOURCE_BRANCH_WIKI_PATH, RESOURCE_BRANCH_XING_PATH
from .base import BaseEvaluation
from mx_crm.models import XingCompanyDb, WikipediaDb, engine, Company, session


class BranchEvaluationLevel(BaseEvaluation):

    def get_excel_data_branches(self, filename):
        """
        Read branch information from the Resources folder
        The branch ratings are used to determine what
        final Wikipedia and XING Branch rating the company
        should get
        """
        list_of_imported_info = self.read_xls(filename)
        dict_of_branches, dict_of_new_squirrel_industries = dict(), dict()

        list_of_imported_info.next()
        for element in list_of_imported_info:
            # print type(element[0].value)
            # print element[0].value
            escaped_branch = ''
            if not element[0].value is None:
                escaped_branch = re.escape(element[0].value.strip())

            # escaped_branch = re.escape(str(element[0].value).encode('utf-8').strip())

            inner_branch_level = int(element[1].value or 0)
            branch = element[0].value.strip() if element[0].value else ''
            mark = element[2].value or 0

            dict_of_branches.setdefault(inner_branch_level, [])
            dict_of_branches[inner_branch_level].append(escaped_branch)

            dict_of_new_squirrel_industries.setdefault(branch, [])
            dict_of_new_squirrel_industries[branch].append(mark)

        for key, value in dict_of_branches.iteritems():
            dict_of_branches[key] = "|".join(value)

        return dict_of_branches, dict_of_new_squirrel_industries

    def calc(self, table='wiki', companies=[]):
        """
        Calculate what branch level the company falls in
        Branch levels are calculated according to the
        branch weights extracted from the Resources branch
        excel files.
        """

        if table == "xing":
            query = self._xing_data(
                fields=[XingCompanyDb.industry_xing],
                filters=[XingCompanyDb.industry_xing!=None, XingCompanyDb.industry_xing!='']
            )
            wiki_branch_filename = settings.RESOURCE_BRANCH_XING_PATH
        elif table == "wiki":
            query = self._wiki_data(
                fields=[WikipediaDb.branch_wikipedia_w],
                filters=[WikipediaDb.branch_wikipedia_w!=None, WikipediaDb.branch_wikipedia_w!='']
            )
            wiki_branch_filename = settings.RESOURCE_BRANCH_WIKI_PATH

        branch_level_dict, new_squirrel_industry_dict = self.get_excel_data_branches(wiki_branch_filename)
        sorted_branch_level_dict = OrderedDict(sorted(branch_level_dict.items(), reverse=False))

        wiki_dict_id, wiki_dict_name = dict(), dict()
        for element in query:
            branch_name = element[2].strip() if element[2] else ''
            wiki_dict_id[element[1]] = branch_name
            wiki_dict_name[element[0].strip().lower() if element[0] else ''] = branch_name

        inner_branch_dict_id_wiki, inner_branch_dict_name_wiki = dict(), dict()

        for element in self._company_data(companies=companies):
            c_name, c_id = element[0].lower().strip(), element[1]
            # if company id from main company table exists in wikipedia table
            for key, value in sorted_branch_level_dict.iteritems():
                if c_id in wiki_dict_id and wiki_dict_id[c_id].strip():
                    # iterate through each level in branch level excel table starting from lowest to highest
                    # compile regex
                    current_level_regex = r'(.*?)(' + value + r')(.*?)'
                    current_level_regex_comp = regex.compile(current_level_regex, re.IGNORECASE)
                    # match current level branch with current branch from main company table
                    match_current_level = current_level_regex_comp.match(wiki_dict_id[c_id])
                    # add final branch level
                    if match_current_level:
                        inner_branch_dict_id_wiki[c_id] = key

                # if company name from main company table exists in wikipedia table
                if c_name in wiki_dict_name and wiki_dict_name[c_name].strip():
                    # iterate through each level in branch level excel table starting from lowest to highest
                    current_level_regex = r'(.*?)(' + value + r')(.*?)'
                    current_level_regex_comp = regex.compile(current_level_regex, re.IGNORECASE)
                    # match current level branch with current branch from main company table
                    match_current_level = current_level_regex_comp.match(wiki_dict_name[c_name])
                    # add final branch level
                    if match_current_level:
                        inner_branch_dict_name_wiki[c_name] = key

        return inner_branch_dict_id_wiki, inner_branch_dict_name_wiki

    def protection_calc_wiki(self, company_name):
        company = company_name
        # print company
        wiki_branch_filename = RESOURCE_BRANCH_WIKI_PATH
        point = 0
        query = session.query(WikipediaDb.branch_wikipedia_w).filter(
            WikipediaDb.company_name_w == company,
        )
        try:
            industry = query[0][0]
        except IndexError as e:
            return 0
        if industry is None:
            # print "None"
            return 0
        # print industry
        wb = load_workbook(wiki_branch_filename)
        sheet = wb.get_sheet_by_name('Tabelle1')
        for row in sheet.values:
            if industry in row:
                values = row[1]
                if values is not None:
                    point = values
        return int(point)

    def protection_calc_xing(self, company_name):
        company = company_name
        xing_branch_filename = RESOURCE_BRANCH_XING_PATH
        point = 0
        query = session.query(XingCompanyDb.industry_xing).filter(
            XingCompanyDb.company_name_x == company,
        )
        try:
            industry = query[0][0]
        except IndexError as e:
            return 0
        if industry is None:
            return 0
        wb = load_workbook(xing_branch_filename)
        sheet = wb.get_sheet_by_name('Tabelle1')
        for row in sheet.values:
            if industry in row:
                values = row[1]
                if values is not None:
                    point = values
        return int(point)

if __name__ == '__main__':
    BEL = BranchEvaluationLevel()
    BEL.calc()
