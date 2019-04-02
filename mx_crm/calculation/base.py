from openpyxl import load_workbook
from sqlalchemy.orm import Session

from mx_crm.models import XingCompanyDb, WikipediaDb, engine, Company


class BaseEvaluation(object):

    session = None

    def __init__(self):
        self.session = Session(engine)

    def __del__(self):
        self.session.close()

    def read_xls(self, excel_file):
        """
        Read city names from the Resources folder
        The city names are used to determine what
        country the company is from to improve
        the location level calculation
        """
        workbook = load_workbook(filename=excel_file)
        worksheet = workbook[workbook.sheetnames[0]]

        for i, row in enumerate(worksheet.rows):
            yield row

    def _attach_to_query(self, query, filters=[], limit=0):
        if filters:
            query = query.filter(*filters)
        if limit:
            query = query.limit(limit)
        return query

    def _xing_data(self, fields=[], filters=[], limit=0):
        query = self.session.query(XingCompanyDb.company_name_x, XingCompanyDb.xc_id, *fields)
        query = self._attach_to_query(query, filters, limit)
        return query.yield_per(5000)

    def _wiki_data(self, fields=[], filters=[], limit=0):
        query = self.session.query(WikipediaDb.company_name_w, WikipediaDb.wc_id, *fields)
        query = self._attach_to_query(query, filters, limit)
        return query.yield_per(5000)

    def _company_data(self, companies=[], limit=0):
        query = self.session.query(
            Company.name, Company.id, Company.website,
            Company.impressum_name, Company.cleaned_name
        )
        if companies:
            query = query.filter(Company.name.in_(companies))
        query = self._attach_to_query(query, limit=limit)
        return query.yield_per(5000)
