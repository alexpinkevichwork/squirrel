import logging

from openpyxl import load_workbook

from mx_crm import settings
from mx_crm.models import session, XingCompanyDb, WikipediaDb

logs_file = settings.rel('mx_crm', 'logs', settings.RESOURCE_SYNC_LOGS)
try:
    logging.basicConfig(filename=logs_file,level=logging.INFO)
except IOError:
    logging.basicConfig(filename="C:\\Users\\admin\\PycharmProjects\\SquirrelRunnerNew\\mx_crm\\logs\\resource_sync_logs.log",
                        level=logging.INFO)



class ResourceSync(object):

    BLACKLISTED_NAMES = ['-']

    wiki_list_name = 'Tabelle1'
    xing_list_name = 'Tabelle1'

    resource_xing_path = settings.RESOURCE_BRANCH_XING_PATH
    resource_wiki_path = settings.RESOURCE_BRANCH_WIKI_PATH

    def __init__(self, *args, **kwargs):
        try:
            self.wiki_wb = load_workbook(filename=self.resource_wiki_path)
            self.xing_wb = load_workbook(filename=self.resource_xing_path)
        except IOError:
            self.wiki_wb = {self.wiki_list_name: None}
            self.xing_wb = {self.wiki_list_name: None}

        self.wiki_ws = self.wiki_wb[self.wiki_list_name]
        self.xing_ws = self.xing_wb[self.xing_list_name]

    def execute_db_data(self, fields, distinct):
        logging.info('Execute industries from DB')
        return dict(
            (i[0].lower(), i[0])
            for i in session.query(*fields).distinct(distinct).all()
            if i and i[0]
        )

    def _read_sheet(self, sheet):
        logging.info('Reading work sheet')
        data = {}
        for row_i, row in enumerate(sheet.rows):
            data[row_i] = row
        return (data, row_i+1)

    def wiki_sync(self):
        if not self.wiki_ws:
            return
        logging.info('WIKIPEDIA sync started')
        resource_data, last_row = self._read_sheet(self.wiki_ws)
        db_data = self.execute_db_data((WikipediaDb.branch_wikipedia_w,), WikipediaDb.branch_wikipedia_w)

        current_file_list = self._build_name_list(resource_data)

        self._determine_and_write_new(db_data, current_file_list, last_row, self.wiki_ws)
        self.wiki_wb.save(self.resource_wiki_path)
        logging.info('WIKIPEDIA sync finished')

    def xing_sync(self):
        if not self.xing_ws:
            return
        logging.info('XING sync started')
        resource_data, last_row = self._read_sheet(self.xing_ws)
        db_data = self.execute_db_data((XingCompanyDb.industry_xing,), XingCompanyDb.industry_xing)

        current_file_list = self._build_name_list(resource_data)

        self._determine_and_write_new(db_data, current_file_list, last_row, self.xing_ws)
        self.xing_wb.save(self.resource_xing_path)
        logging.info('XING sync finished')

    def sync_all(self):
        logging.info('Sync ALL')
        self.wiki_sync()
        self.xing_sync()

    def _build_name_list(self, resource_data):
        return dict(
            (item[0].value.lower(), key)
            for key, item in resource_data.items()
            if item and item[0] and item[0].value
        )

    def _determine_and_write_new(self, db_data, current_file_list, last_row, work_sheet):
        logging.info('Matching industries. Determining new industries.')
        added = 0
        for key_lower, key in db_data.items()[1:]:
            if not key_lower or key_lower in current_file_list or key_lower in self.BLACKLISTED_NAMES:
                continue
            current_row = last_row + added + 1
            work_sheet.cell(row=current_row, column=1).value = key
            #work_sheet.cell(row=current_row, column=2).value = 0
            added += 1


if __name__ == '__main__':
    RS = ResourceSync()
    RS.sync_all()
