# -*- coding: utf-8 -*-
from __future__ import print_function, absolute_import

import re
import sys
import time
from pprint import pprint

import psutil
import logging
import requests
import datetime
import traceback

from subprocess import Popen
from ipwhois import IPWhois
from ipwhois.utils import get_countries
from sqlalchemy import func, or_

from mx_crm import settings
from mx_crm.models import GoogleAnalyticsVisits, WikipediaDb, XingCompanyDb, session, Company
from mx_crm.settings import SCRAPYD_SCHEDULE_URL

logger = logging.getLogger(__name__)


def get_whois(ip_address):
    """
    Function to get the information from RIPE. Sends an IP Address and gets Company name behind it.
    Also gets an address and country of origin. Later we only allow companies from Germany, Switzerland, Austria
    and Netherlands.
    """
    # search the RIPE Database for the given IP
    res = IPWhois(ip_address).lookup_rdap(rate_limit_timeout=30)

    # get the country of the IP Address
    countries = get_countries()
    net = res['network']
    country = countries[net['country']]
    company_name = net['name']

    # get the name of the company behind the IP Address
    # clean the names of the company from Provider description
    # since sometimes it is saved in RIPE as 'Provider Name + Company Name'
    for regex, n in settings.COMPANY_REGEXES:
        m = re.match(regex, company_name)
        if m:
            company_name = m.group(n)
            break

    logger.debug('Country: {} / Possible Entity Name: {}'.format(country, company_name))

    # get the Address of the Company; in rare cases this is an actual company name and not the
    # one found in the description field
    try:
        full_address = res['objects'][res['entities'][0]]['contact']['address'][0]['value']
    except KeyError:
        full_address = ''

    address = full_address.split('\n')[0]
    full_address = full_address.replace('\n', ' ')

    return country, company_name, address, full_address


def sqlalchemy_to_dict(query_result=None):
    cover_dict = {key: getattr(query_result, key) for key in query_result.keys()}
    return cover_dict


def parse_revenue(revenue):
    revenue = revenue.replace("'", '').replace(',', '.')
    revenue = re.split(r'[( \[]', revenue)
    multiplier = 0
    amount = None
    currency = None
    for piece in revenue:
        try:
            amount = float(piece)
            continue
        except ValueError:
            pass

        piece = re.sub('\.$', '', piece.lower())
        if piece in settings.MIO:
            multiplier = 1
            continue
        elif piece in settings.MRD:
            multiplier = 1000
            continue
        elif piece.startswith('eur') or piece == u'€':
            currency = 'Euro'
            continue
        elif 'us' in piece or piece == u'$':
            currency = 'US-Dollar'
            continue
        elif piece in settings.CURRENCIES:
            currency = piece.upper()

    if not multiplier:
        multiplier = .000001

    if not (amount and currency):
        return None, None

    revenue = amount * multiplier
    return revenue, currency


def get_scrapyd_jobs(project_name):
    return requests.get(settings.SCRAPYD_JOBS_URL, {'project': project_name}).json()


def ip_digits(ip_address):
    # Returns IP address without the last three digits
    ip = ip_address.split('.')
    if len(ip) == 4 and ip[-1] != '0':
        ip[-1] = '0'
        return '.'.join(ip)
    return ip_address


def prepare_company_name_for_match(company_name):
    return re.sub('( GmbH$| AG$| mbH$)', '', company_name, flags=re.IGNORECASE).strip().lower()


def convert_to_float(obj):
    try:
        obj = float(obj)
    except (TypeError, ValueError):
        pass
    return obj


def convert_to_int(obj):
    try:
        obj = int(obj)
    except (TypeError, ValueError):
        pass
    return obj


def prepare_date_to_drupal_execute(days, **kwargs):
    def replace_seconds(time):
        splited_time = time.split(':')
        if len(splited_time) > 2:
            return ':'.join(splited_time[:2])
        return time

    if kwargs.get('current_date') and kwargs.get('last_date'):
        current_date = ' '.join([kwargs.get('current_date'), replace_seconds(kwargs.get('current_time', ''))])
        last_date = ' '.join([kwargs.get('last_date'), replace_seconds(kwargs.get('last_time', ''))])

        current_date = datetime.datetime.strptime(current_date, '%Y-%m-%d %H:%M')
        last_date = datetime.datetime.strptime(last_date, '%Y-%m-%d %H:%M')
    else:
        current_date = datetime.now()
        last_date = current_date - datetime.timedelta(days=int(days))

    return (current_date, last_date)


def start_scrapyd(**kwargs):
    logger = get_logger(**kwargs)
    scrapyd = Popen('scrapyd')
    scrapyd_process = psutil.Process(scrapyd.pid)
    logger.info("Scrapyd started")
    time.sleep(5)
    return (scrapyd, scrapyd_process)


def kill_scrapyd(scrapyd, scrapyd_process, **kwargs):
    logger = get_logger(**kwargs)

    logger.info('Killing scrapyd process...')
    try:
        child_proceses = scrapyd_process.children(recursive=True)
    except psutil.NoSuchProcess:
        child_proceses = []
    for child in child_proceses:
        try:
            child.kill()
        except psutil.NoSuchProcess:
            continue
    try:
        scrapyd.kill()
    except psutil.NoSuchProcess:
        pass
    logger.info("Scrapyd killed")


def run_scrapy_process(project_name, spider_name, post_data):
    requests.post(SCRAPYD_SCHEDULE_URL, post_data)
    while True:
        resp = get_scrapyd_jobs(project_name)
        if len(resp['pending']) or len(resp['running']):
            logger.debug('{} spider still working'.format(spider_name))
            time.sleep(5)
        else:
            time.sleep(10)
            break


def print_traceback():
    exc_type, exc_value, exc_traceback = sys.exc_info()
    traceback.print_tb(exc_traceback, limit=1, file=sys.stdout)
    traceback.print_exception(exc_type, exc_value, exc_traceback,
                              limit=2, file=sys.stdout)


def get_logger(**kwargs):
    return kwargs.get('log') or logger


def kill_all_scrappy():
    import os
    cd = os.system('curl http://localhost:6800/listjobs.json?project=default > kill_job.text')
    file = open('kill_job.text', 'r')
    import ast
    a = ast.literal_eval(file.read())
    b = a.values()
    c = b[3]
    for i in c:
        kill = 'curl http://localhost:6800/cancel.json -d project=default -d job={}'.format(i['id'])
        os.system(kill)


def get_scrappy_processes_count():
    import os
    cd = os.system('curl http://localhost:6800/listjobs.json?project=default > kill_job.text')
    file = open('kill_job.text', 'r')
    import ast
    a = ast.literal_eval(file.read())
    b = a.values()
    c = b[3]
    pprint(len(c))


def get_zero_employees_xing():
    from mx_crm.models import session
    from mx_crm.models import XingCompanyDb
    query = session.query(XingCompanyDb.company_name_x).filter(XingCompanyDb.employees_size_xing == 0)
    count = 0
    names_to_update = []
    for i in query:
        count += 1
        names_to_update.append(i[0])
    pprint(count)
    for i in names_to_update[:1000]:
        query = session.query(XingCompanyDb).filter(
            XingCompanyDb.company_name_x == i,
        )
        query.update({XingCompanyDb.manual_entry: "old"}, synchronize_session="fetch")
        session.commit()


def convert_old_to_manul_xing():
    from mx_crm.models import session
    from mx_crm.models import XingCompanyDb
    query = session.query(XingCompanyDb.company_name_x).filter(XingCompanyDb.manual_entry == 'Yes')
    count = 0
    names_to_update = []
    for i in query:
        count += 1
        names_to_update.append(i[0])
    pprint(count)
    for i in names_to_update:
        query = session.query(XingCompanyDb).filter(
            XingCompanyDb.company_name_x == i,
        )
        query.update({XingCompanyDb.manual_entry: "No"}, synchronize_session="fetch")
        session.commit()


def converting_wiki_companies_state():
    from mx_crm.models import session
    from mx_crm.models import WikipediaDb
    query = session.query(WikipediaDb.company_name_w).filter(WikipediaDb.manual_entry == 'old')
    count = 0
    names_to_update = []
    for i in query:
        count += 1
        names_to_update.append(i[0])
    pprint(count)
    for i in names_to_update:
        query = session.query(WikipediaDb).filter(
            WikipediaDb.company_name_w == i,
        )
        query.update({WikipediaDb.manual_entry: "No"}, synchronize_session="fetch")
        session.commit()


def google_analytics_update():
    import subprocess
    from pprint import pprint
    from openpyxl import load_workbook
    from mx_crm.models import session
    from mx_crm.models import Company

    subprocess.call('dir T:\google_analytics > T:\google_analytics\path.txt', shell=True)
    fp = open("T:\google_analytics\path.txt")
    file_path = ''
    for i, line in enumerate(fp):
        if i == 7:
            file_path = line[36:84]
    fp.close()
    file_path = 'T:\google_analytics\{}'.format(file_path)
    split_date_str = file_path[55:-5]
    year = int(split_date_str[:4])
    month = int(split_date_str[4:6])
    day = int(split_date_str[6:])
    date = datetime.datetime(year=year, month=month, day=day)
    wb = load_workbook(file_path)
    sheet = wb.get_sheet_by_name('Datensatz1')
    companies = []
    rows_count = []
    pages = []

    for row in sheet.values:
        print(row)
        name = row[0]
        if name not in companies:
            companies.append(name)
    to_create = []
    for company in companies[1:-1]:
        print(company)
        query = session.query(Company).filter(Company.name == company)
        try:
            a = query[0]
        except IndexError:
            to_create.append(company)
            continue
    # uncomment for creating
    for company in to_create:
        print(company)
        new_company = Company(name=company, timestamp=func.now())
        query = session.add(new_company)
        session.commit()

    # for company in companies[1:2]:  # change to to_create
    #     for row in sheet.values:
    #         if row[0] is company:
    #             print(row[0])
    #             print(row[1])
    #             print(row[7])

    for company in companies[1:-1]:  # change to to_create
        for row in sheet.values:
            if row[0] == company:
                query_c = session.query(Company.id).filter(Company.name == company)
                new_google_analytic_data = GoogleAnalyticsVisits(
                    c_id=query_c[0][0],
                    company_name_g=company,
                    visited_page=row[1],
                    duration=row[7],
                    visit_date=date
                )
                query_g = session.add(new_google_analytic_data)
                session.commit()


def add_google_analytics_accounts_to_report_file(path_to_xl, start_time, end_time):
    from openpyxl import load_workbook
    from mx_crm.models import session
    from mx_crm.models import Company
    # path = 'report{}'.format(path_to_xl[62:102])
    path = 'C:\\Users\\admin\\PycharmProjects\\SquirrelRunnerNew\\mx_crm\\reports'
    from os.path import join
    from os import listdir
    a = [join(path, f) for f in listdir(path)]
    path = a[-1]
    wb = load_workbook(path)
    # wb = load_workbook('C:\\Users\\admin\\PycharmProjects\\SquirrelRunnerNew\\mx_crm\\reports\\{}'.format(path_to_xl[64:]))
    sheet = wb.get_sheet_by_name('Report')
    max_row = sheet.max_row
    names_indexes = {}
    counter = 0
    google_c = []
    google_analytics_companies = session.query(GoogleAnalyticsVisits).filter(
        start_time <= GoogleAnalyticsVisits.visit_date).filter(
        GoogleAnalyticsVisits.visit_date <= end_time).order_by(GoogleAnalyticsVisits.company_name_g)
    for google in google_analytics_companies:
    # for google, company in google_analytics_companies.iteritems():
        company = session.query(Company).filter(Company.name == google.company_name_g).first()
        try:
            wiki_company = session.query(WikipediaDb).filter(WikipediaDb.company_name_w == google.company_name_g).first()
        except IndexError:
            wiki_company = None
        except UnboundLocalError:
            wiki_company = None
        except TypeError:
            wiki_company = None

        try:
            xing_company = session.query(XingCompanyDb).filter(XingCompanyDb.company_name_x == google.company_name_g).first()
        except IndexError:
            xing_company = None
        max_row += 1
        webinterface_link = "http://192.168.0.141:8000/squirrel/accounts/search/{}/".format(google.company_name_g)
        row_A = 'A{}'.format(max_row)       # COMPANY
        sheet[row_A] = '=HYPERLINK("{}", "{}")'.format(webinterface_link, google.company_name_g)

        row_B = 'B{}'.format(max_row)       # WEBSITE MANUAL
        sheet[row_B] = company.manual_entry

        row_C = 'C{}'.format(max_row)       # WEBSITE
        row_D = 'D{}'.format(max_row)       # SESSION LENGTH
        duration = google.duration
        duration_datetime = datetime.datetime(2000,1,1,0,0,0)
        duration_datetime = duration_datetime + datetime.timedelta(seconds=duration)
        hours = duration_datetime.hour
        minutes = duration_datetime.minute
        seconds = duration_datetime.second
        if int(hours) < 10:
            hours = '0{}'.format(hours)
        if int(minutes) < 10:
            minutes = '0{}'.format(minutes)
        if int(seconds) < 10:
            seconds = '0{}'.format(seconds)
        result_time = '{}:{}:{}'.format(hours, minutes, seconds)
        sheet[row_D] = result_time
        row_E = 'E{}'.format(max_row)       # TOTAL SESSION LENGTH
        total_session = session.query(GoogleAnalyticsVisits.duration).filter(
            GoogleAnalyticsVisits.company_name_g == google.company_name_g
        )
        total = 0
        visit_count = 0
        for ts in total_session:
            total = total + ts[0]
            visit_count += 1
        duration_datetime_total = datetime.datetime(2000,1,1,0,0,0)
        duration_datetime_total = duration_datetime_total + datetime.timedelta(seconds=total)
        total_hours = duration_datetime_total.hour
        total_minutes = duration_datetime_total.minute
        total_seconds = duration_datetime_total.second
        if int(total_hours) < 10:
            total_hours = '0{}'.format(total_hours)
        if int(total_minutes) < 10:
            total_minutes = '0{}'.format(total_minutes)
        if int(total_seconds) < 10:
            total_seconds = '0{}'.format(total_seconds)
        result_time_total = '{}:{}:{}'.format(total_hours, total_minutes, total_seconds)
        sheet[row_E] = result_time_total

        row_F = 'F{}'.format(max_row)       # SQUIRREL RATING
        sheet[row_F] = company.squirrel_rating

        row_I = 'I{}'.format(max_row)       # REFER
        page = google.visited_page
        visited_page = 'https://www.mobilexag.de{}'.format(page.encode('utf-8'))
        sheet[row_I] = visited_page

        row_J = 'J{}'.format(max_row)       # Access time.
        sheet[row_J] = google.visit_date

        row_AA = 'AA{}'.format(max_row)     # ACCOUNT_MANUAL
        sheet[row_AA] = company.manual_account_id
        row_AB = 'AB{}'.format(max_row)     # ACCOUNT
        sheet[row_AB] = company.d_crm_account_name
        row_AC = 'AC{}'.format(max_row)     # ACCOUNT OWNER
        sheet[row_AC] = company.d_crm_account_owner
        row_AE = 'AE{}'.format(max_row)     # LONG URL
        sheet[row_AE] = company.website_long
        row_AF = 'AF{}'.format(max_row)     # CLOSED ACTIVITY TYPE
        sheet[row_AF] = company.d_crm_closed_activity_type
        row_AG = 'AG{}'.format(max_row)     # CLOSED ACTIVITY DATE
        sheet[row_AG] = company.d_crm_closed_date
        row_AH = 'AH{}'.format(max_row)     # OPEN ACTIVITY TYPE
        sheet[row_AH] = company.d_crm_open_activity_type
        row_AI = 'AI{}'.format(max_row)     # SCHEDULED DATE
        sheet[row_AI] = company.d_crm_schedule_date
        row_AJ = 'AJ{}'.format(max_row)     # ACCOUNT ID
        sheet[row_AJ] = company.account_id
        row_AK = 'AK{}'.format(max_row)     # RELATIONSHIP TYPE
        sheet[row_AK] = company.d_crm_relationship_type
        row_AM = 'AM{}'.format(max_row)     # ABC
        sheet[row_AM] = company.d_crm_abc_rating

        row_AN = 'AN{}'.format(max_row)     # TOTAL SESSION LENGTH
        sheet[row_AN] = result_time_total
        row_AO = 'AO{}'.format(max_row)     # TOTAL VISITED COUNT
        sheet[row_AO] = visit_count
        row_AS = 'AS{}'.format(max_row)     # SQUIRREL RATING
        sheet[row_AS] = company.squirrel_rating
        row_AT = 'AT{}'.format(max_row)     # LOCATION LEVEL
        sheet[row_AT] = company.mx_crm_location_level
        row_AU = 'AU{}'.format(max_row)     # BRANCH
        sheet[row_AU] = company.mx_crm_branch_level
        row_AV = 'AV{}'.format(max_row)     # GOOGLE EVALUATION
        sheet[row_AV] = company.mx_crm_google_evaluation
        row_AW = 'AW{}'.format(max_row)     # WIKI RATING POINTS
        sheet[row_AW] = company.mx_crm_wiki_rating_points
        row_AX = 'AX{}'.format(max_row)     # XING RATING POINTS
        sheet[row_AX] = company.mx_crm_xing_rating_points
        row_AY = 'AY{}'.format(max_row)     # REVENUE
        sheet[row_AY] = company.mx_crm_revenue_level

        if wiki_company:
            row_L = 'L{}'.format(max_row)    # WIKIPEDIA MANUAL
            sheet[row_L] = wiki_company.manual_entry
            row_M = 'M{}'.format(max_row)    # WIKIPEDIA PAGE
            sheet[row_M] = wiki_company.wiki_url_w
            row_N = 'N{}'.format(max_row)    # WIKIPEDIA REVENUE
            sheet[row_N] = wiki_company.revenue_wikipedia_w
            row_O = 'O{}'.format(max_row)    # REVENUE CURRECY
            sheet[row_O] = wiki_company.revenue_currency_wiki_w
            row_P = 'P{}'.format(max_row)    # EMPLOYEES NUMBER
            sheet[row_P] = wiki_company.employees_wikipedia_w
            row_Q = 'Q{}'.format(max_row)    # WIKIPEDIA CATEGORIES
            sheet[row_Q] = wiki_company.categories_wikipedia_w
            row_R = 'R{}'.format(max_row)    # WIKIPEDIA BRANCHE
            sheet[row_R] = wiki_company.branch_wikipedia_w
            row_S = 'S{}'.format(max_row)    # WIKIPEDIA SUMMARY
            sheet[row_S] = wiki_company.summary_wikipedia_w
            row_AQ = 'AQ{}'.format(max_row)    # WIKIPEDIA MANUAL
            sheet[row_AQ] = wiki_company.manual_entry

        if xing_company is not None:
            print(xing_company)
            row_T = 'T{}'.format(max_row)    # MANUAL ENTRY
            sheet[row_T] = xing_company.manual_entry
            row_U = 'U{}'.format(max_row)    # XING COMPANY PROFILE
            sheet[row_U] = xing_company.xing_url
            row_V = 'V{}'.format(max_row)    # XING COMPANY COUNTRY
            sheet[row_V] = xing_company.country_xing
            row_W = 'W{}'.format(max_row)    # XING COMPANY EMPLOYEES
            sheet[row_W] = xing_company.employees_group_xing_x
            row_X = 'X{}'.format(max_row)    # XING COMPANY EMPLOYEES NUMBER
            sheet[row_X] = xing_company.employees_size_xing
            row_Y = 'Y{}'.format(max_row)    # XING DESCRIPTION
            sheet[row_Y] = xing_company.description_xing
            row_Z = 'Z{}'.format(max_row)    # XING INDUSTRY
            sheet[row_Z] = xing_company.industry_xing
            row_AR = 'AR{}'.format(max_row)  # MANUAL ENTRY
            sheet[row_AR] = xing_company.manual_entry

        counter += 1
        # if not sheet.row_dimensions[max_row].collapsed:
        #     sheet.row_dimensions[max_row].hidden = True
        #     sheet.row_dimensions[max_row].outlineLevel = 1
        if sheet[max_row][0].value == sheet[max_row - 1][0].value:
            sheet.row_dimensions[max_row].hidden = True
            # sheet.row_dimensions[max_row - 1].hidden = True
            sheet.row_dimensions[max_row].outlineLevel = 1
            # sheet.row_dimensions[max_row - 1].outlineLevel = 1

    # for company in google_analytics_companies.values():
    #     indexes = []
    #     for index, row in enumerate(sheet.values):
    #         finded_regex = re.findall(r'\"(.+?)\"', row[0])
    #         for i in finded_regex:
    #             if i == company.name:
    #                 indexes.append(index)
    #         # for i in indexes:
    #         #     sheet.row_dimensions[i].hidden = True
    #         #     sheet.row_dimensions[i].outlineLevel = 1
    #     names_indexes[company.name] = indexes
    # for name, index in names_indexes.iteritems():
    #     for i in index:
    #         sheet.row_dimensions[i].hidden = True
    wb.save(path)
    return 1


def force_update_google_analytics_companies(companies):
    updated_names = set()
    for company in companies:
        name = u'update_{}'.format(company.company_name_g)
        updated_names.add(name)
    return updated_names


def update_report_file_with_google_analytics(file_path):
    return


def select_empty_wiki_link_companys():
    from pprint import pprint
    companies = session.query(Company.name, Company.wikipedia_url).filter(
        Company.wikipedia_url == ''
    )
    wiki_objects = session.query(WikipediaDb.company_name_w, WikipediaDb.wiki_url_w).filter(
        WikipediaDb.company_name_w.in_([j[0] for j in companies]), WikipediaDb.wiki_url_w != ""
    )
    for obj in wiki_objects:
        if obj[1] and obj[1] != "N/A":
            company = companies.filter_by(name=obj[0])
            print(company[0])
            print(obj[1])
            company.update({Company.wikipedia_url: obj[1]}, synchronize_session="fetch")
            session.commit()


def select_empty_wiki_link_wiki_db():
    wiki_objects = session.query(WikipediaDb.company_name_w, WikipediaDb.wiki_url_w).filter(or_(
        WikipediaDb.wiki_url_w == "N/A",
        WikipediaDb.wiki_url_w == "",
        WikipediaDb.wiki_url_w == None
    ))
    companies = session.query(Company.name, Company.wikipedia_url).filter(
        Company.name.in_([j[0] for j in wiki_objects]),
        or_(
            Company.wikipedia_url != "N/A",
            Company.wikipedia_url != "",
            Company.wikipedia_url != None,
            # Company.wikipedia_url != u""
        )
    )
    # print(companies.count())
    for i in companies:
        if i[1] and i[1] != 'N/A':
            wiki_obj = wiki_objects.filter_by(company_name_w=i[0])
            print(wiki_obj[0][0])
            print(i[1])
            # wiki_obj.update({WikipediaDb.wiki_url_w: i[1]}, synchronize_session="fetch")
            # session.commit()
