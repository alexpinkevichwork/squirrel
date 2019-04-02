# -*- coding: utf-8 -*-
from __future__ import print_function, absolute_import

import re

from datetime import datetime, timedelta
from pprint import pprint

from openpyxl import load_workbook

from sqlalchemy.sql import func
from sqlalchemy.sql.expression import between

from mx_crm import settings
from mx_crm.models import session as alchemy_session, Company, DbIpDatabase, Accesslog, session
from mx_crm.exporter import create_report
from mx_crm.utils import ip_digits, prepare_company_name_for_match
from mx_crm.queries import Access, DrupalSession, CompanyEntry
from mx_crm.class_helpers import AdditionalFields

import logging
logger = logging.getLogger(__name__)


class ReportMatche(object):

    list_name = 'AllAccountsLastOpenAndClosedDat'   # 'Accounts Last Activity_Date_Web'
    accounts_file_path = settings.ACCOUNTS_LIST_PATH

    def __init__(self, accounts_file_path='', *args, **kwargs):
        if accounts_file_path:
            self.accounts_file_path = accounts_file_path
        try:
            self.accounts_file = load_workbook(filename=self.accounts_file_path, read_only=False)
        except IOError:
            raise Exception("Accounts file - NOT FOUND !!!")

        if kwargs.get('list_name'):
            self.list_name = kwargs.get('list_name')

        self.ws = self.accounts_file[self.list_name]

    def _get_account_headers(self, rows):
        rows.next()
        return tuple([c.value for c in rows.next()])

    def read_account_file_urls(self, names):
        rows = self.ws.rows
        data = {}
        data1 = {}
        columns_names = self._get_account_headers(rows)
        for row_i, row in enumerate(rows):
            obj = {}
            obj1 = {}
            if row_i < 2:
                continue
            elif row:
                count = 3
                for cell_i, cell in enumerate(row):
                    obj[columns_names[cell_i]] = cell.value
                    obj1[columns_names[cell_i]] = self.ws.cell(row=count, column=1).hyperlink.target
                    count += 1
                company_name = obj['Account'] or ''
                if company_name.lower() == names[0]:
                    obj[u'hyperlink'] = obj1.values()[0]
                    data[company_name] = obj
        return data, data1

    def read_account_file_websites_ids(self, company_names):
        ids = []
        websites = []
        websites_ids_dict = {}
        data = {}
        returned_data = {}
        rows = self.ws.rows
        columns_names = self._get_account_headers(rows)
        for name in company_names:
            query = session.query(Company.name, Company.account_id).filter(Company.name == name)
            try:
                ids.append(query[0][1])
            except IndexError:
                continue
        for name in company_names:
            query = session.query(Company.name, Company.website).filter(Company.name == name)
            try:
                websites.append(query[0][1])
            except IndexError:
                continue
        websites_ids_dict = dict(zip(websites, ids))
        for website, id in websites_ids_dict.iteritems():
            try:
                website = re.sub('www\d?\.', '', website).rstrip('/').lower()
                website = re.sub('http(s)?\:\/\/', '', website)
            except TypeError:
                continue
            if id != u'':
                for row_i, row in enumerate(rows):
                    obj = {}
                    if row_i < 2:
                        continue
                    elif row:
                        for cell_i, cell in enumerate(row):
                            obj[columns_names[cell_i]] = cell.value
                        company_name = obj['Account'] or ''
                        company_name = prepare_company_name_for_match(company_name)
                        p_account_id = obj['Account ID'] or ''
                        if p_account_id and p_account_id == id:
                            data[website] = obj
                            data[company_name] = obj
                        elif company_name in company_names:
                            data[company_name] = obj
                            data[website] = obj
            else:
                for row_i, row in enumerate(rows):
                    obj = {}
                    if row_i < 2:
                        continue
                    elif row:
                        for cell_i, cell in enumerate(row):
                            obj[columns_names[cell_i]] = cell.value
                        company_name = obj['Account'] or ''
                        p_website = obj['Web Site Url'] or ''
                        p_website = re.sub('www\d?\.', '', p_website).rstrip('/').lower()
                        p_website = re.sub('http(s)?\:\/\/', '', p_website)
                        p_website_rev = self._reverse_de_com(p_website)
                        if p_website and p_website == website:
                            data[p_website] = obj
                        elif p_website and p_website_rev == website:
                            data[p_website_rev] = obj
                        elif company_name in company_names:
                            data[company_name] = obj
        return (data, columns_names)

    def read_account_file(self, websites, company_names):
        data = {}

        rows = self.ws.rows
        columns_names = self._get_account_headers(rows)
        for row_i, row in enumerate(rows):
            obj = {}
            if row_i < 2:
                continue
            elif row:
                for cell_i, cell in enumerate(row):
                    obj[columns_names[cell_i]] = cell.value
                company_name = obj['Account'] or ''
                company_name = prepare_company_name_for_match(company_name)
                p_website = obj['Web Site Url'] or ''
                p_website = re.sub('www\d?\.', '', p_website).rstrip('/').lower()
                p_website = re.sub('http(s)?\:\/\/', '', p_website)
                p_website_rev = self._reverse_de_com(p_website)
                if p_website and p_website in websites:
                    data[p_website] = obj
                elif p_website and p_website_rev in websites:
                    data[p_website_rev] = obj
                elif company_name in company_names:
                    data[company_name] = obj
        return (data, columns_names)

    def read_account_file_with_id(self, company_names):
        data = {}
        ids = []
        rows = self.ws.rows
        columns_names = self._get_account_headers(rows)
        for name in company_names:
            query = session.query(Company.name, Company.account_id).filter(Company.name == name)
            try:
                ids.append(query[0][1])
            except IndexError:
                continue
        for row_i, row in enumerate(rows):
            obj = {}
            if row_i < 2:
                continue
            elif row:
                for cell_i, cell in enumerate(row):
                    obj[columns_names[cell_i]] = cell.value
                company_name = obj['Account'] or ''
                company_name = prepare_company_name_for_match(company_name)
                p_account_id = obj['Account ID'] or ''
                if p_account_id and p_account_id in ids:
                    data[p_account_id] = obj
                elif company_name in company_names:
                    data[company_name] = obj
        return data

    def _reverse_de_com(self, website):
        if re.search('.de$', website):
            return re.sub('\.de$', '.com', website)
        elif re.search('.com$', website):
            return re.sub('\.com$', '.de', website)
        return website


class MatchExecutor(object):

    def _build_entry_companies(self):
        end_time = datetime.now()
        start_time = end_time - timedelta(days=settings.REPORT_PERIOD_IN_DAYS)

        accesslog = [
            Access(*res) for res in alchemy_session.query(
                Accesslog.timestamp, Accesslog.hostname, Accesslog.path, Accesslog.url, Accesslog.title).filter(
                between(Accesslog.timestamp, func.unix_timestamp(start_time), func.unix_timestamp(end_time)),
                Accesslog.title != 'Generate image style',
                Accesslog.hostname.notin_(settings.IPS_BLACKLIST)
            ).order_by(Accesslog.hostname, Accesslog.timestamp)]

        blacklist = {tup[0].lower() for tup in alchemy_session.query(Company.name).filter(
            Company.type_main.in_(['Blacklist', 'Spam', 'Provider']))}

        ips_info = {tup[0]: tup[1:] for tup in alchemy_session.query(
            DbIpDatabase.ip_ip, DbIpDatabase.ip_country, DbIpDatabase.ip_name,
            DbIpDatabase.ip_name_2, DbIpDatabase.ip_address
        )}

        res = {}
        drupal_session = DrupalSession()
        session_length = 0
        len_accesslog = len(accesslog[:-1]) - 1
        for index, request in enumerate(accesslog[:-1]):
            host = ip_digits(request.hostname)
            access_datetime = datetime.fromtimestamp(int(request.timestamp))

            next_request = accesslog[index + 1]
            next_request_host = ip_digits(next_request.hostname)
            next_request_access_datetime = datetime.fromtimestamp(int(next_request.timestamp))

            difference = next_request_access_datetime - access_datetime

            is_continue = False
            if host == next_request_host and difference.seconds < settings.MAXIMUM_DIFFERENCE_BETWEEN_SESSIONS.seconds:
                session_length += difference.seconds
                is_continue = True
            elif host == next_request_host:
                session_length += settings.LONG_SESSION_DEFAULT
                is_continue = True
            elif host != next_request_host:
                session_length += settings.LONG_SESSION_DEFAULT

            if index and host == ip_digits(accesslog[index - 1].hostname) and host != next_request_host:
                drupal_session.append(request)
            elif host == next_request_host:
                drupal_session.append(request)
                is_continue = True

            if is_continue and index != len_accesslog:
                continue

            if host in ips_info:
                country, company_name, address_result, full_address_result = ips_info[host]

            company_name = company_name.lower()

            if company_name and country in settings.RELEVANT_COUNTRIES \
                    and company_name not in settings.PROVIDERS_BLACKLIST \
                    and company_name not in blacklist \
                    and not any(word in company_name for word in settings.COMPANIES_BLACKLIST) \
                    and not any(re.search(regexp, company_name) for regexp in settings.PROVIDERS_BLACKLIST_REGEXPS) \
                    and not any(re.search(regexp, company_name) for regexp in settings.COMPANIES_BLACKLIST_REGEXPS):

                if company_name not in res:
                    res[company_name] = CompanyEntry(*ips_info[host], sessions=[])

                res[company_name].sessions.append(drupal_session)
                res[company_name].session_length = timedelta(seconds=session_length)

            session_length = 0
            drupal_session = DrupalSession()

        return res

    def _get_companies(self, entry_companies_keys):
        return alchemy_session.query(Company). \
                   filter(
                        Company.name.in_(entry_companies_keys) &
                        (Company.website != None) &
                        (Company.website != 'NA')
                    ).all()

    def _get_and_prepare_websites(self, companies_entries):
        return set(
            re.sub('www\d?\.', '', company.website).rstrip('/').lower()
            for company in companies_entries if company.website
        )

    def _get_websites_from_objects(self, companies_entries):
        return set(
            company.website for company in companies_entries if company.website
        )

    def _get_and_prepare_company_names(self, companies_entries):
        return set(
            prepare_company_name_for_match(company.name)
            for company in companies_entries
        )

    def _get_and_prepare_company_names_from_objects(self, companie_entries):
        return set(
            company.name for company in companie_entries
        )

    def create_report(self, entry_companies=[], google_analytics_companies={}, dates={}):
        report_worker = ReportMatche()#'/home/vladimir/Загрузки/Accounts Last Activity_Date_WebSiteURL_LongURL.xlsx')
        dates_next=dates
        entry_companies = entry_companies or self._build_entry_companies()
        logger.info('entry_companies')
        logger.info(entry_companies)
        # query = session.query(Company).filter(Company.name == 'tdg fuer roto frank ag')
        # entry_companies.update({'tdg fuer roto frank ag': query[0]})
        # logger.info(query)
        logger.info('entry_companies')
        logger.info(entry_companies)
        logger.info('type(entry_companies)')
        logger.info(type(entry_companies))
        companies_entries = self._get_companies(entry_companies.keys())
        # websites = self._get_websites_from_objects(companies_entries)
        websites = self._get_and_prepare_websites(companies_entries)
        company_names = self._get_and_prepare_company_names_from_objects(companies_entries)
        data_links = {}
        for name in company_names:
            k, l = ReportMatche().read_account_file_urls([name])
            try:
                try:
                    link = k.values()[0].get(u'hyperlink')
                    data_links[name] = link
                except IndexError:
                    link = ''
                    data_links[name] = link
                    # pprint(k.values()[0].get(u'hyperlink'))
            except AttributeError:
                link = ''
                data_links[name] = link
        af = AdditionalFields()
        logger.info('af')
        logger.info(af)
        account_data = {}
        total_fields = af.total_fields(entry_companies.keys())
        logger.info('total_fields')
        logger.info(total_fields)
        for name in company_names:
            account_data_name, account_headers = report_worker.read_account_file_websites_ids([name])
            account_data.update(account_data_name)
        # a_d = report_worker.read_account_file_with_id(company_names)
        create_report(entry_companies, account_data, account_headers, total_fields, data_links, google_analytics_companies,
                      dates=dates_next)
        logger.info('create_report')
        logger.info(create_report)


if __name__ == '__main__':
    me = MatchExecutor()
    me.create_report()
